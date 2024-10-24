from openpyxl import Workbook
from openpyxl import load_workbook

excel_fn = "/tmp/1.xlsx"
wb = load_workbook(filename=excel_fn)

stns = wb.sheetnames
st = wb['abc']
# st = wb.active

vcf_str = ""
cnt = 1
for row in st.rows:
    c1 = row[0].value
    c2 = row[1].value
    c3 = row[2].value
    c4 = row[3].value
    tels = str(row[4].value)
    name = "CF"
    if c4:
        name = name + c4
    print(name)
    note = ""
    if c1:
        note += c1
    if c3:
        note += c3
    if cnt > 3000:
        break
    tel = ""
    for i in tels.split(";"):
        if i:
            tel += f'TEL;CELL:{i}\n'
            print(i)
    vcf_str += f'''BEGIN:VCARD
N;CHARSET=UTF-8:{name}
{tel}NOTE;CHARSET=UTF-8:{note}
END:VCARD'''
    cnt = cnt+1

with open("/tmp/1.vcf", "w", encoding="utf-8") as f:
    f.write(vcf_str)
